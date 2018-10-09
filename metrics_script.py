from openpyxl import load_workbook, Workbook   # downloaded from pip
import os
import glob
import datetime
import boto3
from botocore.exceptions import ClientError
from botocore.config import Config


class Metrics:

    def __init__(self):
        self.metrics = [('Applications', 'Task Description', 'Task Classification',	'Assigned Date', 'Completion Date', 'Effort Hours',	'Status', 'Assigned To', 'Secondary Index')]
        self.vacation_dates = [('Application Name', 'Employee', 'From Date', 'To Date', 'No of Days', 'Vacation Type', 'Comment')]
        self.time_sheet = [('Resource Name', 'Apllication', 'Total hours in catw')]
        self.application = 'Not mentioned'


# Getting all metrics with date validations
    def get_metrics(self, ws, employee):
        for i in range(4, ws.max_row+1):
            application_name = ws.cell(row=i, column=1).value
            if application_name is None:
                break
            self.application = application_name
            task_description = ws.cell(row=i, column=2).value
            task_classification = ws.cell(row=i, column=3).value
            assigndate = str(ws.cell(row=i, column=4).value)
            try:
                assigndate = datetime.datetime.strptime(assigndate, '%Y-%m-%d %H:%M:%S').date()
            except ValueError:
                raise ValueError(f"Incorrect assign date format {assigndate} for {employee} in metrics sheet")
            completiondate = str(ws.cell(row=i, column=5).value)
            try:
                completiondate = datetime.datetime.strptime(completiondate, '%Y-%m-%d %H:%M:%S').date()
            except ValueError:
                raise ValueError(f"Incorrect completion date format {completiondate} for {employee} in metrics sheet")
            effortHours = ws.cell(row=i, column=6).value
            status = ws.cell(row=i, column=7).value
            if status is None:
                status = 'Closed'
            secondary_index = 'Un known'
            if task_classification == 'Application Training' or task_classification == 'Documentation' or task_classification == 'On Call Abend Resolution' or task_classification == 'Research Work' or task_classification == 'Verification':
                secondary_index = 'Functional Support'
            elif task_classification is 'Program Changes /Testing  /Design':
                secondary_index = 'CR'
            else:
                secondary_index = 'Processing Support'
            if secondary_index == 'Un known':
                raise ValueError(f'Cant find the secondary index for {employee} for {task_classification} in metrics sheet')
            self.metrics.append((application_name, task_description, task_classification, assigndate, completiondate, effortHours, status, employee, secondary_index))


# Getting all vacation dates with date validations
    def get_vacation_dates(self, ws, employee):
        for i in range(2, ws.max_row+1):
            name = employee
            if ws.cell(row=i, column=1).value is None:
                break
            from_date = str(ws.cell(row=i, column=2).value)
            try:
                from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d %H:%M:%S').date()
            except ValueError:
                raise ValueError(f"Incorrect from date format {from_date} for {employee} in vacation dates")
            to_date = str(ws.cell(row=i, column=3).value)
            try:
                to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d %H:%M:%S').date()
            except ValueError:
                raise ValueError(f"Incorrect to date format {to_date} for {employee} in vacation dates")
            count_leaves = self.get_no_of_leaves(from_date.day, to_date.day + 1)
            vacation_type = ws.cell(row=i, column=4).value
            comment = ws.cell(row=i, column=5).value
            if vacation_type is None:
                vacation_type = 'Not Mentioned'
            if comment is None:
                comment = 'Not Mentioned'
            self.vacation_dates.append((self.application, name, from_date, to_date, count_leaves, vacation_type, comment))


# Getting all time sheet with date validations
    def get_time_sheet(self, ws, employee):
        name = employee
        catw_hours = ws.cell(row=2, column=2).value
        if catw_hours is None:
            catw_hours = 0
        if type(catw_hours) is not int:
            raise TypeError(f"Incorrect catw hours format for {employee} in time sheet")
        self.time_sheet.append((name, self.application, catw_hours))


# getting no of days between leaves
    def get_no_of_leaves(self, start, end):
        no_of_leaves = 0
        # sundays and saturdays and public holidays in this month
        holidays = [1, 2, 8, 9, 13, 15, 16, 22, 23, 29, 30]
        for x in range(start, end):
            if x not in holidays:
                no_of_leaves = no_of_leaves + 1
        return no_of_leaves


# final save to excel
    def save_data_to_excel(self):
        wb = Workbook()
        ws = wb.active
        for row in self.metrics:
            ws.append(row)
        wb.save("Metrics.xlsx")
        wb = Workbook()
        ws = wb.active
        for row in self.vacation_dates:
            ws.append(row)
        wb.save("VacationDates.xlsx")
        wb = Workbook()
        ws = wb.active
        for row in self.time_sheet:
            ws.append(row)
        wb.save("CatwTimeSheet.xlsx")

# Sending Email
    def send_email(self, to_address, body_text):
        AWS_ACCESS_KEY = 'AKIAJWJCXFVULQ5QJO4Q'
        AWS_SECRET_KEY = '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'
        SENDER = "tharundintakurthi@gmail.com"
        RECIPIENT = to_address
        AWS_REGION = "us-east-1"
        SUBJECT = "Metrics Data Error in Share point please rectify it"
        BODY_TEXT = body_text
        CHARSET = "UTF-8"
        client = boto3.client('ses', region_name=AWS_REGION, aws_access_key_id = AWS_ACCESS_KEY, aws_secret_access_key = AWS_SECRET_KEY,config=Config(proxies={'https': '16.153.99.11:8088'}))
        try:
            client.send_email(
                Destination={
                    'ToAddresses': [
                        RECIPIENT,
                    ],
                },
                Message={
                    'Body': {
                        'Text': {
                            'Charset': CHARSET,
                            'Data': BODY_TEXT,
                        },
                    },
                    'Subject': {
                        'Charset': CHARSET,
                        'Data': SUBJECT,
                    },
                },
                Source=SENDER,
            )
        except ClientError as e:
            print(e.response['Error']['Message'])


# starting of project
    def start_Project(self):
        path = 'C:\\Users\\hpadmin\\Desktop\\aug'
        extension = 'xlsx'
        os.chdir(path)
        result = [i for i in glob.glob('*.{}'.format(extension))]
        for employee in result:
            excel_path = path + '\\' + employee
            work_book = load_workbook(excel_path)
            sheets_list = work_book.sheetnames
            employee = employee[:(len(employee)-5)]
            self.get_metrics(work_book[sheets_list[0]], employee)
            self.get_vacation_dates(work_book['Vacation Dates'], employee)
            self.get_time_sheet(work_book['Timesheet'], employee)
        self.save_data_to_excel()
        # print(self.metrics)
        # print(self.vacation_dates)
        # print(self.time_sheet)


a = Metrics()
a.start_Project()
