from openpyxl import load_workbook   # downloaded from pip
import os
import glob
import datetime


# Getting all metrics with date validations
def get_metrics(ws, employee):
    for i in range(4, ws.max_row):
        app = ws.cell(row=i, column=1).value
        if app is None:
            break
        td = ws.cell(row=i, column=2).value
        ts = ws.cell(row=i, column=3).value
        assigndate = str(ws.cell(row=i, column=4).value)
        try:
            assigndate = datetime.datetime.strptime(assigndate, '%Y-%m-%d %H:%M:%S').date()
        except ValueError:
            raise ValueError(f"Incorrect assign date format {assigndate} for {employee} in metrics sheet")
        completiondate = str(ws.cell(row=i, column=5).value)
        try:
            completiondate = datetime.datetime.strptime(completiondate, '%Y-%m-%d %H:%M:%S').date()
        except ValueError:
            raise ValueError(f"Incorrect completion date format {completiondate} for {employee} in metrics sheet")
        effortHours = ws.cell(row=i, column=6).value
        status = ws.cell(row=i, column=7).value
        metrics.append((app, td, ts, assigndate, completiondate, effortHours, status, employee))


# Getting all vacation dates with date validations
def get_vacation_dates(ws, employee):
    for i in range(2, ws.max_row):
        name = employee
        if ws.cell(row=i, column=1).value is None:
            break
        from_date = str(ws.cell(row=i, column=2).value)
        try:
            from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d %H:%M:%S').date()
        except ValueError:
            raise ValueError(f"Incorrect from date format {from_date} for {employee} in vacation dates")
        to_date = str(ws.cell(row=i, column=3).value)
        try:
            to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d %H:%M:%S').date()
        except ValueError:
            raise ValueError(f"Incorrect to date format {to_date} for {employee} in vacation dates")
        count_leaves = get_no_of_leaves(from_date.day, to_date.day + 1)
        vacation_type = ws.cell(row=i, column=4).value
        comment = ws.cell(row=i, column=5).value
        if vacation_type is None:
            vacation_type = 'Not Mentioned'
        if comment is None:
            comment = 'Not Mentioned'
        vacation_dates.append((name, from_date, to_date, count_leaves, vacation_type, comment))


# Getting all time sheet with date validations
def get_time_sheet(ws, employee):
    name = employee
    catw_hours = ws.cell(row=2, column=2).value
    if catw_hours is None:
        catw_hours = 0
    if type(catw_hours) is not int:
        raise TypeError(f"Incorrect catw hours format for {employee} in time sheet")
    time_sheet.append((name, catw_hours))


# getting no of days between leaves
def get_no_of_leaves(start, end):
    no_of_leaves = 0
    # sundays and saturdays and public holidays in this month
    holidays = [20, 21]
    for x in range(start, end):
        if x not in holidays:
            no_of_leaves = no_of_leaves + 1
    return no_of_leaves


path = 'C:\\Users\\hpadmin\\Desktop\\august'
extension = 'xlsx'
os.chdir(path)
metrics = []
vacation_dates = []
time_sheet = []
result = [i for i in glob.glob('*.{}'.format(extension))]
for employee in result:
    excel_path = path + '\\' + employee
    work_book = load_workbook(excel_path)
    sheets_list = work_book.sheetnames
    get_metrics(work_book[sheets_list[0]], employee)
    get_vacation_dates(work_book['Vacation Dates'], employee)
    get_time_sheet(work_book['Timesheet'], employee)
# print(metrics)
print(vacation_dates)
# print(len(time_sheet))
