from openpyxl import load_workbook
import datetime

metrics = []

wb2 = load_workbook('b.xlsx')
print(wb2.sheetnames)
# Metrics Sheet Code
ws = wb2['Metrics-2018']
for i in range(4, ws.max_row):
    app = ws.cell(row=i, column=1).value
    td = ws.cell(row=i, column=2).value
    ts = ws.cell(row=i, column=3).value
    assigndate = str(ws.cell(row=i, column=4).value)
    try:
        assigndate = datetime.datetime.strptime(assigndate, '%Y-%m-%d %H:%M:%S').date()
    except ValueError:
        raise ValueError(f"Incorrect assign date format {assigndate}")
    completiondate = str(ws.cell(row=i, column=5).value)
    try:
        completiondate = datetime.datetime.strptime(completiondate, '%Y-%m-%d %H:%M:%S').date()
    except ValueError:
        raise ValueError(f"Incorrect completion date format {completiondate}")
    effortHours = ws.cell(row=i, column=6).value
    status = ws.cell(row=i, column=7).value
    if app is None:
        break
    else:
        metrics.append((app, td, ts, assigndate, completiondate, effortHours, status))
print(metrics)
